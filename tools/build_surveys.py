#!/usr/bin/env python3
"""Generate three sample XLSForms for ODK Central, plus their media CSVs.

Each form demonstrates one of the mapping patterns in
docs/ODK-Central-Integration.md:

    training_event.xlsx     distinct_entity  - roster of attendees, sex-split
    proposal_inclusion.xlsx ratio_fields     - numerator and denominator
    variety_trial.xlsx      ratio_fields + weight, and a second distinct_entity

Run:  python3 tools/build_surveys.py     (needs openpyxl)
Output: surveys/*.xlsx and surveys/media/*.csv
"""
import csv
import pathlib
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("openpyxl is required:  pip install openpyxl", file=sys.stderr)
    raise SystemExit(1)

ROOT = pathlib.Path(__file__).resolve().parent.parent
OUT = ROOT / "surveys"
MEDIA = OUT / "media"

HEADER_FILL = PatternFill("solid", fgColor="DDEBF7")
GROUP_FILL = PatternFill("solid", fgColor="F2F2F2")


def write_sheet(ws, header, rows, widths):
    ws.append(header)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = HEADER_FILL
        c.alignment = Alignment(vertical="center")
    for r in rows:
        ws.append([r.get(h, "") for h in header])
    for i, wdt in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = wdt
    ws.freeze_panes = "A2"


def build(path, survey, choices, settings, survey_header=None):
    wb = Workbook()
    sh = wb.active
    sh.title = "survey"
    header = survey_header or ["type", "name", "label", "hint", "required",
                               "relevant", "constraint", "constraint_message",
                               "calculation", "appearance", "readonly"]
    write_sheet(sh, header, survey, [34, 26, 52, 40, 10, 34, 30, 34, 30, 18, 10])

    ch = wb.create_sheet("choices")
    write_sheet(ch, ["list_name", "name", "label"], choices, [24, 18, 44])

    st = wb.create_sheet("settings")
    write_sheet(st, ["form_title", "form_id", "version", "instance_name"],
                [settings], [46, 26, 14, 34])

    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Media CSVs.
#
# In production these are published by ODK Central from an Entity List, which
# Central serves to the form under the same filename. Shipping static CSVs here
# means the forms can be loaded and tested before Entity Lists exist; the form
# itself does not change when you switch over.
# ---------------------------------------------------------------------------

def write_media():
    MEDIA.mkdir(parents=True, exist_ok=True)

    partners = [("DP-%03d" % n, "Delivery Partner %03d" % n,
                 ["NARS", "Scaling", "Private"][n % 3],
                 ["NG", "GH", "KE", "TZ", "UG", "ZM"][n % 6])
                for n in range(1, 41)]
    with (MEDIA / "delivery_partners.csv").open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["name", "label", "partner_type", "country"])
        w.writerows(partners)

    scientists = [("NARS-SCI-%04d" % n, "Scientist %04d" % n,
                   ["NG", "GH", "KE", "TZ", "UG", "ZM"][n % 6])
                  for n in range(1, 61)]
    with (MEDIA / "national_scientists.csv").open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["name", "label", "country"])
        w.writerows(scientists)

    varieties = [("GAZ-%03d" % n, "Variety %03d" % n,
                  ["CASSAVA", "YAM", "COWPEA", "SOYBEAN", "MAIZE", "BANANA"][n % 6])
                 for n in range(1, 31)]
    with (MEDIA / "varieties.csv").open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["name", "label", "crop"])
        w.writerows(varieties)


# ---------------------------------------------------------------------------
# Shared choice lists
# ---------------------------------------------------------------------------

SEX = [
    {"list_name": "sex", "name": "F", "label": "Female"},
    {"list_name": "sex", "name": "M", "label": "Male"},
    {"list_name": "sex", "name": "other", "label": "Other / prefer not to say"},
]
YESNO = [
    {"list_name": "yesno", "name": "yes", "label": "Yes"},
    {"list_name": "yesno", "name": "no", "label": "No"},
]
COUNTRIES = [
    {"list_name": "country", "name": c, "label": n} for c, n in [
        ("NG", "Nigeria"), ("BJ", "Benin"), ("GH", "Ghana"),
        ("CD", "Democratic Republic of Congo"), ("CM", "Cameroon"),
        ("KE", "Kenya"), ("TZ", "Tanzania"), ("UG", "Uganda"),
        ("RW", "Rwanda"), ("ZM", "Zambia"), ("MW", "Malawi"), ("MZ", "Mozambique")]
]
CROPS = [
    {"list_name": "crop", "name": c, "label": n} for c, n in [
        ("CASSAVA", "Cassava"), ("YAM", "Yam"), ("BANANA", "Banana and plantain"),
        ("COWPEA", "Cowpea"), ("SOYBEAN", "Soybean"), ("MAIZE", "Maize")]
]
META = [
    {"type": "start", "name": "start"},
    {"type": "end", "name": "end"},
    {"type": "today", "name": "today"},
    {"type": "deviceid", "name": "deviceid"},
]


# ---------------------------------------------------------------------------
# Form 1 - Training event.  Pattern: distinct_entity
#
# Feeds KPI 6 (delivery partners trained) and KPI 7 (national scientists
# trained), both distinct counts, both sex-disaggregated.
#
# The point of this form: there is NO "how many attended" question and NO
# "percentage female" question. Both are derived from the roster, so they cannot
# disagree with it.
# ---------------------------------------------------------------------------

def form_training():
    survey = META + [
        {"type": "begin_group", "name": "event", "label": "Training event"},
        {"type": "text", "name": "event_title", "label": "Title of the training",
         "required": "yes"},
        {"type": "date", "name": "event_date", "label": "Date the training ended",
         "required": "yes",
         "constraint": ". <= today()",
         "constraint_message": "The training cannot end in the future."},
        {"type": "select_one country", "name": "country", "label": "Country",
         "required": "yes",
         "hint": "Coded to ISO 3166. Free-text locations are rejected at ingestion."},
        {"type": "select_one crop", "name": "crop",
         "label": "Primary commodity (if any)"},
        {"type": "integer", "name": "training_days", "label": "Duration in days",
         "constraint": ". > 0 and . < 120"},
        {"type": "end_group", "name": "event"},

        # --- Roster 1: delivery partners (KPI 6) -----------------------------
        {"type": "select_one yesno", "name": "has_partners",
         "label": "Were delivery partner organisations trained?", "required": "yes"},
        {"type": "begin_repeat", "name": "partner_attendee",
         "label": "Delivery partner attendees",
         "relevant": "${has_partners} = 'yes'",
         "hint": "One row per attending organisation. Do not enter a total."},
        {"type": "select_one_from_file delivery_partners.csv", "name": "partner_id",
         "label": "Partner organisation", "required": "yes",
         "appearance": "search",
         "hint": "Select from the register. Never type a new name here."},
        {"type": "select_one sex", "name": "partner_sex",
         "label": "Sex of the person who attended", "required": "yes"},
        {"type": "select_one yesno", "name": "partner_completed",
         "label": "Completed the full course?", "required": "yes",
         "hint": "Only completions count toward the KPI."},
        {"type": "end_repeat", "name": "partner_attendee"},

        # --- Roster 2: national scientists (KPI 7) ---------------------------
        {"type": "select_one yesno", "name": "has_scientists",
         "label": "Did national scientists attend?", "required": "yes"},
        {"type": "begin_repeat", "name": "scientist_attendee",
         "label": "National scientist attendees",
         "relevant": "${has_scientists} = 'yes'"},
        {"type": "select_one_from_file national_scientists.csv", "name": "scientist_id",
         "label": "Scientist", "required": "yes", "appearance": "search"},
        {"type": "select_one sex", "name": "scientist_sex",
         "label": "Sex", "required": "yes"},
        {"type": "select_one yesno", "name": "scientist_completed",
         "label": "Completed the full course?", "required": "yes"},
        {"type": "end_repeat", "name": "scientist_attendee"},

        # --- Derived, for the enumerator only --------------------------------
        {"type": "calculate", "name": "n_partners",
         "calculation": "count(${partner_attendee})"},
        {"type": "calculate", "name": "n_scientists",
         "calculation": "count(${scientist_attendee})"},
        {"type": "note", "name": "roster_summary",
         "label": ("Recorded: ${n_partners} partner attendees, "
                   "${n_scientists} scientist attendees. "
                   "These totals are derived from the rosters and are not submitted "
                   "as indicator values."),
         "readonly": "yes"},

        {"type": "image", "name": "attendance_sheet",
         "label": "Photograph of the signed attendance sheet",
         "hint": "Attached as evidence against the observation."},
    ]
    choices = SEX + YESNO + COUNTRIES + CROPS
    settings = {"form_title": "Training event (standard module)",
                "form_id": "training_event",
                "version": "2025.3",
                "instance_name": "concat(${event_title}, ' - ', ${event_date})"}
    return build(OUT / "training_event.xlsx", survey, choices, settings)


# ---------------------------------------------------------------------------
# Form 2 - Proposal inclusion screening.  Pattern: ratio_fields
#
# Feeds KPI 18, 19 and 20, all percentages. The form collects a roster of
# proposals and the flags on each; the numerator and denominator are counted
# from it. No percentage is ever typed in.
# ---------------------------------------------------------------------------

def form_proposals():
    survey = META + [
        {"type": "begin_group", "name": "period", "label": "Reporting period"},
        {"type": "date", "name": "reporting_date",
         "label": "Date of this screening return", "required": "yes"},
        {"type": "select_one country", "name": "country", "label": "Lead country"},
        {"type": "end_group", "name": "period"},

        {"type": "begin_repeat", "name": "proposal",
         "label": "Approved proposals in this period",
         "hint": "One row per approved proposal. The denominator is the number of rows."},
        {"type": "text", "name": "proposal_ref", "label": "Proposal reference",
         "required": "yes"},
        {"type": "text", "name": "proposal_title", "label": "Title"},
        {"type": "date", "name": "approval_date", "label": "Date approved",
         "required": "yes"},
        {"type": "select_one yesno", "name": "has_gender_youth",
         "label": "Includes a gender / youth / social inclusion component?",
         "required": "yes"},
        {"type": "select_one yesno", "name": "has_climate",
         "label": "Includes adaptation / mitigation activities?", "required": "yes"},
        {"type": "select_one yesno", "name": "has_nutrition",
         "label": "Includes nutrition and health activities?", "required": "yes"},
        {"type": "calculate", "name": "flag_inclusion",
         "calculation": "if(${has_gender_youth} = 'yes', 1, 0)"},
        {"type": "calculate", "name": "flag_climate",
         "calculation": "if(${has_climate} = 'yes', 1, 0)"},
        {"type": "calculate", "name": "flag_nutrition",
         "calculation": "if(${has_nutrition} = 'yes', 1, 0)"},
        {"type": "end_repeat", "name": "proposal"},

        {"type": "calculate", "name": "proposals_approved_total",
         "calculation": "count(${proposal})"},
        {"type": "calculate", "name": "proposals_with_inclusion",
         "calculation": "sum(${flag_inclusion})"},
        {"type": "calculate", "name": "proposals_with_climate",
         "calculation": "sum(${flag_climate})"},
        {"type": "calculate", "name": "proposals_with_nutrition",
         "calculation": "sum(${flag_nutrition})"},

        {"type": "note", "name": "totals_note",
         "label": ("${proposals_approved_total} approved proposals recorded. "
                   "Shares are calculated by the KPI engine from the numerator "
                   "and denominator - do not enter a percentage anywhere."),
         "readonly": "yes"},
    ]
    choices = YESNO + COUNTRIES
    settings = {"form_title": "Proposal inclusion screening",
                "form_id": "proposal_inclusion",
                "version": "2025.1",
                "instance_name": "concat('Proposals - ', ${reporting_date})"}
    return build(OUT / "proposal_inclusion.xlsx", survey, choices, settings)


# ---------------------------------------------------------------------------
# Form 3 - Variety trial and release.  Patterns: ratio_fields + weight,
#                                                and distinct_entity
#
# Feeds KPI 9 (genetic gain, weighted by trial count) and KPI 8 (varieties
# released, de-duplicated by variety).
# ---------------------------------------------------------------------------

def form_variety():
    survey = META + [
        {"type": "begin_group", "name": "assessment", "label": "Trial assessment"},
        {"type": "date", "name": "assessment_date", "label": "Date of assessment",
         "required": "yes"},
        {"type": "select_one country", "name": "country", "label": "Country",
         "required": "yes"},
        {"type": "select_one crop", "name": "crop", "label": "Crop", "required": "yes"},
        {"type": "end_group", "name": "assessment"},

        # --- KPI 9: weighted average ----------------------------------------
        {"type": "begin_group", "name": "gain", "label": "Yield-related genetic gain"},
        {"type": "decimal", "name": "gain_pct",
         "label": "Yield-related genetic gain for this crop (%)", "required": "yes",
         "constraint": ". >= -50 and . <= 100",
         "constraint_message": "Enter a percentage, not a fraction."},
        {"type": "calculate", "name": "gain_basis", "calculation": "100"},
        {"type": "integer", "name": "trial_count",
         "label": "Number of trials or plots this figure is based on",
         "required": "yes",
         "constraint": ". > 0",
         "hint": ("This is the weight. Gains are combined across crops weighted "
                  "by it, never as a simple average.")},
        {"type": "end_group", "name": "gain"},

        # --- KPI 8: distinct count ------------------------------------------
        {"type": "select_one yesno", "name": "any_release",
         "label": "Were any varieties officially released this period?",
         "required": "yes"},
        {"type": "begin_repeat", "name": "release", "label": "Varieties released",
         "relevant": "${any_release} = 'yes'",
         "hint": "One row per variety. A variety released in two countries is still one variety."},
        {"type": "select_one_from_file varieties.csv", "name": "variety_id",
         "label": "Variety", "required": "yes", "appearance": "search"},
        {"type": "date", "name": "release_date", "label": "Date of official release",
         "required": "yes"},
        {"type": "text", "name": "gazette_ref",
         "label": "Release gazette reference", "required": "yes",
         "hint": "KPI 8 requires evidence; a release without this is held at ingestion."},
        {"type": "file", "name": "gazette_document",
         "label": "Gazette document or scan"},
        {"type": "end_repeat", "name": "release"},

        {"type": "calculate", "name": "n_releases", "calculation": "count(${release})"},
        {"type": "note", "name": "release_note",
         "label": ("${n_releases} release(s) recorded. De-duplication across "
                   "projects and programs happens in the KPI engine, not here."),
         "readonly": "yes"},
    ]
    choices = YESNO + COUNTRIES + CROPS
    settings = {"form_title": "Variety trial and release",
                "form_id": "variety_trial",
                "version": "2025.2",
                "instance_name": "concat(${crop}, ' - ', ${assessment_date})"}
    return build(OUT / "variety_trial.xlsx", survey, choices, settings)


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    write_media()
    made = [form_training(), form_proposals(), form_variety()]
    for p in made:
        print("wrote", p.relative_to(ROOT), "(%.1f KB)" % (p.stat().st_size / 1024))
    for p in sorted(MEDIA.glob("*.csv")):
        print("wrote", p.relative_to(ROOT))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
